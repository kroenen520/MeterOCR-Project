#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电表OCR批量识别工具 - 主程序
功能：扫描程序所在目录的所有子文件夹，批量识别电表图片，输出Excel表格
作者：AI Assistant
版本：1.0.0
"""

import os
import sys
import re
import time
import warnings
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Optional, Dict
import argparse

warnings.filterwarnings('ignore')

# 检查依赖
try:
    from paddleocr import PaddleOCR
except ImportError:
    print("错误: 未安装PaddleOCR")
    print("请运行: pip install paddleocr paddlepaddle")
    input("按回车键退出...")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: 未安装openpyxl，将使用CSV格式输出")


class MeterOCRApp:
    """电表OCR应用程序"""
    
    def __init__(self):
        """初始化应用程序"""
        self.ocr = None
        self.station_pattern = re.compile(r'440\d{15}')
        self.current_pattern = re.compile(r'\b\d+\.?\d*\b')
        self.image_extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp'}
        
    def initialize_ocr(self) -> bool:
        """初始化OCR引擎"""
        try:
            print("正在初始化OCR引擎...")
            print("首次运行会自动下载模型文件，请耐心等待...\n")
            
            self.ocr = PaddleOCR(
                use_textline_orientation=True,
                lang='ch'
            )
            print("✓ OCR引擎初始化完成！\n")
            return True
        except Exception as e:
            print(f"✗ OCR引擎初始化失败: {e}")
            return False
    
    def get_subfolders(self, base_path: str) -> List[Path]:
        """
        获取所有子文件夹
        
        Args:
            base_path: 基础路径（程序所在目录）
            
        Returns:
            子文件夹列表
        """
        base = Path(base_path)
        subfolders = [d for d in base.iterdir() if d.is_dir()]
        return sorted(subfolders)
    
    def get_images_in_folder(self, folder_path: str) -> List[Path]:
        """
        获取文件夹内所有图片
        
        Args:
            folder_path: 文件夹路径
            
        Returns:
            图片文件列表
        """
        folder = Path(folder_path)
        images = []
        
        # 只搜索当前文件夹，不递归子文件夹
        for ext in self.image_extensions:
            images.extend(folder.glob(f'*{ext}'))
            images.extend(folder.glob(f'*{ext.upper()}'))
        
        return sorted(images)
    
    def extract_station_code(self, text: str) -> Optional[str]:
        """提取站址编码"""
        cleaned = re.sub(r'\s+', '', text)
        matches = self.station_pattern.findall(cleaned)
        
        if matches:
            return matches[0]
        
        # 宽松匹配（处理OCR误识别）
        loose_pattern = re.compile(r'440[\dO]{15}')
        matches = loose_pattern.findall(cleaned)
        if matches:
            code = matches[0].replace('O', '0').replace('o', '0')
            if len(code) == 18:
                return code
        
        return None
    
    def extract_current_value(self, text: str) -> Optional[str]:
        """提取电流值"""
        lines = text.split('\n')
        candidates = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            numbers = self.current_pattern.findall(line)
            for num_str in numbers:
                try:
                    num = float(num_str)
                    if 0.1 <= num <= 100:
                        candidates.append((num_str, num, line))
                except ValueError:
                    continue
        
        if not candidates:
            return None
        
        # 优先策略
        for num_str, num, line in candidates:
            if any(kw in line for kw in ['电表', '电流', 'A', '安培']):
                return num_str
        
        decimal_candidates = [(n, v) for n, v, _ in candidates if '.' in n]
        if decimal_candidates:
            return decimal_candidates[0][0]
        
        for num_str, num, _ in candidates:
            if 1 <= num <= 50:
                return num_str
        
        return candidates[0][0]
    
    def process_image(self, image_path: str) -> Tuple[Optional[str], Optional[str]]:
        """处理单张图片"""
        try:
            result = self.ocr.ocr(image_path, cls=True)
            
            if not result or not result[0]:
                return None, None
            
            all_text = []
            for line in result[0]:
                if line:
                    all_text.append(line[1][0])
            
            full_text = '\n'.join(all_text)
            station_code = self.extract_station_code(full_text)
            current_value = self.extract_current_value(full_text)
            
            return station_code, current_value
            
        except Exception as e:
            print(f"    处理出错: {e}")
            return None, None
    
    def process_folder(self, folder_path: Path) -> List[Dict]:
        """处理单个文件夹"""
        results = []
        folder_name = folder_path.name
        
        print(f"\n📁 处理文件夹: {folder_name}")
        print("-" * 50)
        
        images = self.get_images_in_folder(str(folder_path))
        
        if not images:
            print(f"  ⚠ 未找到图片文件")
            return results
        
        print(f"  找到 {len(images)} 张图片")
        
        for idx, img_path in enumerate(images, 1):
            print(f"  [{idx}/{len(images)}] {img_path.name}", end=" ")
            
            station_code, current_value = self.process_image(str(img_path))
            
            result = {
                'folder': folder_name,
                'filename': img_path.name,
                'station_code': station_code if station_code else '',
                'current_value': current_value if current_value else '',
                'status': '✓' if (station_code and current_value) else '△' if (station_code or current_value) else '✗'
            }
            results.append(result)
            
            # 显示结果
            info = []
            if station_code:
                info.append(f"站址:{station_code}")
            if current_value:
                info.append(f"电流:{current_value}A")
            
            if info:
                print(f" -> {', '.join(info)}")
            else:
                print(f" -> 未识别")
        
        return results
    
    def generate_output_filename(self) -> str:
        """生成输出文件名（日期时间格式）"""
        now = datetime.now()
        filename = now.strftime("%Y%m%d%H%M%S") + ".xlsx"
        return filename
    
    def save_to_excel(self, results: List[Dict], output_path: str) -> bool:
        """保存结果到Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "识别结果"
            
            # 表头
            headers = ['文件夹', '站址编码', '电流值(A)', '文件名', '识别状态']
            ws.append(headers)
            
            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据
            for result in results:
                ws.append([
                    result['folder'],
                    result['station_code'],
                    result['current_value'],
                    result['filename'],
                    result['status']
                ])
            
            # 列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 12
            
            # 边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"保存Excel失败: {e}")
            return False
    
    def save_to_csv(self, results: List[Dict], output_path: str) -> bool:
        """保存结果到CSV（备用）"""
        try:
            import csv
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['文件夹', '站址编码', '电流值(A)', '文件名', '识别状态'])
                
                for result in results:
                    writer.writerow([
                        result['folder'],
                        result['station_code'],
                        result['current_value'],
                        result['filename'],
                        result['status']
                    ])
            return True
            
        except Exception as e:
            print(f"保存CSV失败: {e}")
            return False
    
    def run(self):
        """运行主程序"""
        # 获取程序所在目录
        if getattr(sys, 'frozen', False):
            # 打包后的EXE运行
            app_dir = Path(sys.executable).parent
        else:
            # 直接运行Python脚本
            app_dir = Path(__file__).parent
        
        print("=" * 60)
        print("  电表OCR批量识别工具 v1.0.0")
        print("=" * 60)
        print(f"程序目录: {app_dir}")
        print("=" * 60)
        
        # 初始化OCR
        if not self.initialize_ocr():
            input("按回车键退出...")
            return
        
        # 获取所有子文件夹
        subfolders = self.get_subfolders(str(app_dir))
        
        if not subfolders:
            print("\n✗ 未找到子文件夹！")
            print("请在程序目录下创建文件夹并放入电表照片。")
            input("按回车键退出...")
            return
        
        print(f"发现 {len(subfolders)} 个子文件夹:")
        for i, folder in enumerate(subfolders, 1):
            print(f"  {i}. {folder.name}")
        
        # 处理所有文件夹
        all_results = []
        for folder in subfolders:
            results = self.process_folder(folder)
            all_results.extend(results)
        
        # 生成输出文件
        if all_results:
            output_filename = self.generate_output_filename()
            output_path = app_dir / output_filename
            
            print("\n" + "=" * 60)
            print("正在保存结果...")
            
            if EXCEL_AVAILABLE:
                if self.save_to_excel(all_results, str(output_path)):
                    print(f"✓ 结果已保存: {output_filename}")
                else:
                    # 尝试CSV
                    csv_path = str(output_path).replace('.xlsx', '.csv')
                    if self.save_to_csv(all_results, csv_path):
                        print(f"✓ 结果已保存: {output_filename.replace('.xlsx', '.csv')}")
            else:
                csv_path = str(output_path).replace('.xlsx', '.csv')
                if self.save_to_csv(all_results, csv_path):
                    print(f"✓ 结果已保存: {output_filename.replace('.xlsx', '.csv')}")
            
            # 统计
            success = sum(1 for r in all_results if r['status'] == '✓')
            partial = sum(1 for r in all_results if r['status'] == '△')
            failed = sum(1 for r in all_results if r['status'] == '✗')
            
            print("\n识别统计:")
            print(f"  总计: {len(all_results)} 张图片")
            print(f"  完全识别: {success} 张")
            print(f"  部分识别: {partial} 张")
            print(f"  识别失败: {failed} 张")
        else:
            print("\n✗ 未找到可识别的图片！")
        
        print("=" * 60)
        input("按回车键退出...")


def main():
    """程序入口"""
    app = MeterOCRApp()
    app.run()


if __name__ == '__main__':
    main()
